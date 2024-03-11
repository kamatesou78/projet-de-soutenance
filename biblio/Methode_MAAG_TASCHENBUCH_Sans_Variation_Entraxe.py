import math
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMessageBox
import openpyxl


class Ui_METHODE_MAAGTASCHENBUCH_SANS_VARIATION_ENTRAXE(object):

    def calculer_engrenage(self):
        try:   
            Z1 = int(self.lineEdit.text())
            Z2 = int(self.lineEdit_2.text()) 
                                    
            X1 = 0.5*((math.log(Z2 / Z1))/(math.log((Z2 * Z1) / 100)))
            X2 = - X1

        except ValueError:
            QMessageBox.warning(None, "Erreur de saisie", "Il est impossible d'effectuer le calcul. Vérifiez les valeurs d'entrée.")
            return
        except ZeroDivisionError:
            QMessageBox.warning(None, "Erreur de saisie", "Il est impossible d'effectuer le calcul. Vérifiez les valeurs d'entrée.")
            return

        # Affichage des resultats dans les LineEdit
        self.lineEdit_3.setText(str(X1))
        self.lineEdit_4.setText(str(X2))
        

    def sauvegarder_dans_excel(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        noms_labels = [
            self.label_3.text(),
            self.label_3.text(),
            self.label_7.text(),
            self.label_7.text()
        ]

        symboles = [
            'Z1', 'Z2',  'X1',  'X2'            
            
        ]


        valeurs = [
            self.lineEdit.text(),
            self.lineEdit_2.text(),
            self.lineEdit_3.text(),
            self.lineEdit_4.text()
       
        ]

        unites = ['', '', '', ''

        ]

        entete = ['Nom(s)', 'Symbole(s)', 'Valeur(s)', 'Unité(s)']
        ws.append(entete)

        for i in range(len(noms_labels)):
            ws.append([noms_labels[i], symboles[i], valeurs[i], unites[i]])

        for row in ws.iter_rows(min_row=1, max_row=5, min_col=1, max_col=4):
            for cell in row:
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(border_style='thin'),
                    right=openpyxl.styles.Side(border_style='thin'),
                    top=openpyxl.styles.Side(border_style='thin'),
                    bottom=openpyxl.styles.Side(border_style='thin')
                )

        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=4):
            for cell in row:
                cell.fill = openpyxl.styles.PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
                cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)


        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(None, "Enregistrer le fichier Excel", "", "Excel Files (*.xlsx)")
        if file_path:
            wb.save(file_path)
            QtWidgets.QMessageBox.information(None, "Enregistrement", "Fichier Excel enregistré avec succès !")



    def setupUi(self, METHODE_MAAGTASCHENBUCH_SANS_VARIATION_ENTRAXE):
        METHODE_MAAGTASCHENBUCH_SANS_VARIATION_ENTRAXE.setObjectName("METHODE_MAAGTASCHENBUCH_SANS_VARIATION_ENTRAXE")
        METHODE_MAAGTASCHENBUCH_SANS_VARIATION_ENTRAXE.resize(1119, 441)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("reducteur_gearexpert.gif"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        METHODE_MAAGTASCHENBUCH_SANS_VARIATION_ENTRAXE.setWindowIcon(icon)
        self.gridLayout_2 = QtWidgets.QGridLayout(METHODE_MAAGTASCHENBUCH_SANS_VARIATION_ENTRAXE)
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.scrollArea = QtWidgets.QScrollArea(METHODE_MAAGTASCHENBUCH_SANS_VARIATION_ENTRAXE)
        self.scrollArea.setWidgetResizable(True)
        self.scrollArea.setObjectName("scrollArea")
        self.scrollAreaWidgetContents = QtWidgets.QWidget()
        self.scrollAreaWidgetContents.setGeometry(QtCore.QRect(0, 0, 1105, 427))
        self.scrollAreaWidgetContents.setObjectName("scrollAreaWidgetContents")
        self.gridLayout = QtWidgets.QGridLayout(self.scrollAreaWidgetContents)
        self.gridLayout.setObjectName("gridLayout")
        self.verticalLayout = QtWidgets.QVBoxLayout()
        self.verticalLayout.setObjectName("verticalLayout")
        self.label = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.verticalLayout.addWidget(self.label)
        self.label_2 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.verticalLayout.addWidget(self.label_2, 0, QtCore.Qt.AlignHCenter)
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.label_3 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.horizontalLayout.addWidget(self.label_3, 0, QtCore.Qt.AlignHCenter)
        self.label_4 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")
        self.horizontalLayout.addWidget(self.label_4, 0, QtCore.Qt.AlignHCenter)
        self.lineEdit = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit.setFont(font)
        self.lineEdit.setObjectName("lineEdit")
        self.horizontalLayout.addWidget(self.lineEdit, 0, QtCore.Qt.AlignHCenter)
        self.label_5 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_5.setFont(font)
        self.label_5.setObjectName("label_5")
        self.horizontalLayout.addWidget(self.label_5, 0, QtCore.Qt.AlignHCenter)
        self.lineEdit_2 = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit_2.setFont(font)
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.horizontalLayout.addWidget(self.lineEdit_2, 0, QtCore.Qt.AlignHCenter)
        self.verticalLayout.addLayout(self.horizontalLayout)
        spacerItem = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.verticalLayout.addItem(spacerItem)
        self.pushButton = QtWidgets.QPushButton(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton.setFont(font)
        self.pushButton.setObjectName("pushButton")

        self.pushButton.clicked.connect(self.calculer_engrenage) # Liaison du clic du bouton à la fonction de calcul
        #################################################################################################

        self.verticalLayout.addWidget(self.pushButton, 0, QtCore.Qt.AlignHCenter)
        spacerItem1 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.verticalLayout.addItem(spacerItem1)
        self.pushButton_1 = QtWidgets.QPushButton(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_1.setFont(font)
        self.pushButton_1.setObjectName("pushButton_1")

        self.pushButton_1.clicked.connect(self.sauvegarder_dans_excel)  # Connecte le clic du bouton à la fonction de sauvegarde
        ###################################################################################
        
        self.verticalLayout.addWidget(self.pushButton_1, 0, QtCore.Qt.AlignHCenter)
        self.label_6 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_6.setFont(font)
        self.label_6.setObjectName("label_6")
        self.verticalLayout.addWidget(self.label_6, 0, QtCore.Qt.AlignHCenter)
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        self.label_7 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_7.setFont(font)
        self.label_7.setObjectName("label_7")
        self.horizontalLayout_2.addWidget(self.label_7, 0, QtCore.Qt.AlignHCenter)
        self.label_8 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_8.setFont(font)
        self.label_8.setObjectName("label_8")
        self.horizontalLayout_2.addWidget(self.label_8, 0, QtCore.Qt.AlignHCenter)
        self.lineEdit_3 = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit_3.setFont(font)
        self.lineEdit_3.setObjectName("lineEdit_3")
        self.horizontalLayout_2.addWidget(self.lineEdit_3, 0, QtCore.Qt.AlignHCenter)
        self.label_9 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_9.setFont(font)
        self.label_9.setObjectName("label_9")
        self.horizontalLayout_2.addWidget(self.label_9, 0, QtCore.Qt.AlignHCenter)
        self.lineEdit_4 = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit_4.setFont(font)
        self.lineEdit_4.setObjectName("lineEdit_4")
        self.horizontalLayout_2.addWidget(self.lineEdit_4, 0, QtCore.Qt.AlignHCenter)
        self.verticalLayout.addLayout(self.horizontalLayout_2)
        spacerItem2 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.verticalLayout.addItem(spacerItem2)
        self.pushButton_2 = QtWidgets.QPushButton(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_2.setFont(font)
        self.pushButton_2.setObjectName("pushButton_2")
        self.verticalLayout.addWidget(self.pushButton_2, 0, QtCore.Qt.AlignHCenter)
        self.gridLayout.addLayout(self.verticalLayout, 0, 0, 1, 1)
        self.scrollArea.setWidget(self.scrollAreaWidgetContents)
        self.gridLayout_2.addWidget(self.scrollArea, 0, 0, 1, 1)

        self.retranslateUi(METHODE_MAAGTASCHENBUCH_SANS_VARIATION_ENTRAXE)
        QtCore.QMetaObject.connectSlotsByName(METHODE_MAAGTASCHENBUCH_SANS_VARIATION_ENTRAXE)

    def retranslateUi(self, METHODE_MAAGTASCHENBUCH_SANS_VARIATION_ENTRAXE):
        _translate = QtCore.QCoreApplication.translate
        METHODE_MAAGTASCHENBUCH_SANS_VARIATION_ENTRAXE.setWindowTitle(_translate("METHODE_MAAGTASCHENBUCH_SANS_VARIATION_ENTRAXE", "METHODE_MAAG-TASCHENBUCH_SANS_VARIATION_ENTRAXE"))
        self.label.setText(_translate("METHODE_MAAGTASCHENBUCH_SANS_VARIATION_ENTRAXE", "     DETERMINATION DES DEPORTS AVEC LA METHODE DE MAAG-TASCHENBUCH\n"
"                                         SANS VARIATION D\'ENTRAXE"))
        self.label_2.setText(_translate("METHODE_MAAGTASCHENBUCH_SANS_VARIATION_ENTRAXE", "Valeurs d\'entrée"))
        self.label_3.setText(_translate("METHODE_MAAGTASCHENBUCH_SANS_VARIATION_ENTRAXE", "Nombre de dent"))
        self.label_4.setText(_translate("METHODE_MAAGTASCHENBUCH_SANS_VARIATION_ENTRAXE", "Z1"))
        self.label_5.setText(_translate("METHODE_MAAGTASCHENBUCH_SANS_VARIATION_ENTRAXE", "Z2"))
        self.pushButton.setText(_translate("METHODE_MAAGTASCHENBUCH_SANS_VARIATION_ENTRAXE", "CALCULEZ"))
        self.pushButton_1.setText(_translate("METHODE_MAAGTASCHENBUCH_SANS_VARIATION_ENTRAXE", "SAUVEGARDER LES RESULTATS DU CALCUL"))
        self.label_6.setText(_translate("METHODE_MAAGTASCHENBUCH_SANS_VARIATION_ENTRAXE", "Valeurs de sortie"))
        self.label_7.setText(_translate("METHODE_MAAGTASCHENBUCH_SANS_VARIATION_ENTRAXE", "Les deports"))
        self.label_8.setText(_translate("METHODE_MAAGTASCHENBUCH_SANS_VARIATION_ENTRAXE", "X1"))
        self.label_9.setText(_translate("METHODE_MAAGTASCHENBUCH_SANS_VARIATION_ENTRAXE", "X2"))
        self.pushButton_2.setText(_translate("METHODE_MAAGTASCHENBUCH_SANS_VARIATION_ENTRAXE", "RETOUR"))


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    METHODE_MAAGTASCHENBUCH_SANS_VARIATION_ENTRAXE = QtWidgets.QWidget()
    ui = Ui_METHODE_MAAGTASCHENBUCH_SANS_VARIATION_ENTRAXE()
    ui.setupUi(METHODE_MAAGTASCHENBUCH_SANS_VARIATION_ENTRAXE)
    METHODE_MAAGTASCHENBUCH_SANS_VARIATION_ENTRAXE.show()
    sys.exit(app.exec_())
