
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMessageBox
import openpyxl
import math 

class Ui_METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE(object):


    def calculer_engrenage(self):
        try:
                
            Z1 = int(self.lineEdit.text())
            Z2 = int(self.lineEdit_2.text()) 
            ANT_F = float(self.lineEdit_3.text())
            A_H = float(self.lineEdit_6.text())
                                    
            # Calcul de M
            M = ((2 * ANT_F) / (Z1 + Z2)) * math.cos(math.radians(A_H))

            # Trouver la valeur la plus proche dans la liste
            liste = [0.06, 0.08, 0.10, 0.12, 0.15, 0.20, 0.25, 0.30, 0.40, 0.50, 0.75, 1, 1.25, 1.5, 2, 2.5, 3, 4, 5, 6, 8, 10, 12, 16, 20, 25, 32, 40, 50, 60, 0.07, 0.09, 0.11, 0.14, 0.18, 0.22, 0.28, 0.35, 0.45, 0.55, 0.7, 0.9, 1.125, 1.375, 1.75, 2.75, 3.5, 4.5, 5.5, 7, 9, 11, 14, 18, 22, 28, 36, 45, 55, 70]
            V_N = min(liste, key=lambda x: abs(x - M))

            # Calcul de A
            A = (V_N /(math.cos(math.radians(A_H)))) * ((Z1 + Z2) / 2)

            Alpha_to = math.atan(math.tan(math.radians(20))/math.cos(math.radians(A_H)))
            Alpha_to_degre = math.degrees(Alpha_to)

            # Calcul d'Alpha_p en degrés
            Alpha_to_p = math.acos((A / ANT_F) * math.cos(math.radians(Alpha_to_degre)))
            Alpha_to_p_degre = math.degrees(Alpha_to_p) 

            # Calcul de inv_Alpha_to_p_degre et inv_Alpha_to
            inv_Alpha_to_p_degre = math.tan(math.radians(Alpha_to_p_degre)) - ((math.pi * Alpha_to_p_degre) / 180)
            inv_Alpha_to = math.tan(math.radians(Alpha_to_degre)) - ((math.pi * Alpha_to_degre) / 180)

            #Calcul de E(somme des deports)

            E = (inv_Alpha_to_p_degre - inv_Alpha_to) * ((Z1 + Z2) / (2 * math.tan(math.radians(20))))

            X1 = 0.5*(E + (1 - E)*((math.log(Z2 / Z1))/(math.log((Z2 * Z1) / 100))))

            X2 = E - X1

        except ValueError:
            QMessageBox.warning(None, "Erreur de saisie", "Il est impossible d'effectuer le calcul. Vérifiez les valeurs d'entrée.")
            return
        except ZeroDivisionError:
            QMessageBox.warning(None, "Erreur de saisie", "Il est impossible d'effectuer le calcul. Vérifiez les valeurs d'entrée.")
            return

        # Affichage des resultats dans les LineEdit
        self.lineEdit_4.setText(str(X1))
        self.lineEdit_5.setText(str(X2))
   

    def sauvegarder_dans_excel(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        noms_labels = [
            self.label_5.text(),
            self.label_5.text(),
            self.label_11.text(),
            self.label_3.text(),
            self.label_7.text(),
            self.label_7.text(),
        ]

        symboles = [
             "Z1", "Z2", "a'", 'β', 'X1', 'X2'    
        ]


        valeurs = [
            self.lineEdit.text(),
            self.lineEdit_2.text(),
            self.lineEdit_3.text(),
            self.lineEdit_6.text(),
            self.lineEdit_4.text(),
            self.lineEdit_5.text(),
        ]

        unites = [ '', '', 'mm', 'degré', '', '' ]

        entete = ['Nom(s)', 'Symbole(s)', 'Valeur(s)', 'Unité(s)']
        ws.append(entete)

        for i in range(len(noms_labels)):
            ws.append([noms_labels[i], symboles[i], valeurs[i], unites[i]])

        for row in ws.iter_rows(min_row=1, max_row=7, min_col=1, max_col=4):
            for cell in row:
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(border_style='thin'),
                    right=openpyxl.styles.Side(border_style='thin'),
                    top=openpyxl.styles.Side(border_style='thin'),
                    bottom=openpyxl.styles.Side(border_style='thin')
                )

        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=4):
            for cell in row:
                cell.fill = openpyxl.styles.PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
                cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)


        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(None, "Enregistrer le fichier Excel", "", "Excel Files (*.xlsx)")
        if file_path:
            wb.save(file_path)
            QtWidgets.QMessageBox.information(None, "Enregistrement", "Fichier Excel enregistré avec succès !")


    def setupUi(self, METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE):
        METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE.setObjectName("METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE")
        METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE.resize(645, 480)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("reducteur_gearexpert.gif"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE.setWindowIcon(icon)
        self.gridLayout_2 = QtWidgets.QGridLayout(METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE)
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.scrollArea = QtWidgets.QScrollArea(METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE)
        self.scrollArea.setWidgetResizable(True)
        self.scrollArea.setObjectName("scrollArea")
        self.scrollAreaWidgetContents = QtWidgets.QWidget()
        self.scrollAreaWidgetContents.setGeometry(QtCore.QRect(0, 0, 795, 453))
        self.scrollAreaWidgetContents.setObjectName("scrollAreaWidgetContents")
        self.gridLayout_3 = QtWidgets.QGridLayout(self.scrollAreaWidgetContents)
        self.gridLayout_3.setObjectName("gridLayout_3")
        self.verticalLayout = QtWidgets.QVBoxLayout()
        self.verticalLayout.setObjectName("verticalLayout")
        self.label = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(20)
        font.setBold(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.verticalLayout.addWidget(self.label, 0, QtCore.Qt.AlignHCenter)
        self.label_2 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.verticalLayout.addWidget(self.label_2, 0, QtCore.Qt.AlignHCenter)
        self.gridLayout = QtWidgets.QGridLayout()
        self.gridLayout.setObjectName("gridLayout")
        self.lineEdit_2 = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit_2.setFont(font)
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.gridLayout.addWidget(self.lineEdit_2, 0, 4, 1, 1, QtCore.Qt.AlignHCenter)
        self.label_47 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_47.setFont(font)
        self.label_47.setObjectName("label_47")
        self.gridLayout.addWidget(self.label_47, 1, 1, 1, 1, QtCore.Qt.AlignHCenter)
        self.label_31 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_31.setFont(font)
        self.label_31.setObjectName("label_31")
        self.gridLayout.addWidget(self.label_31, 0, 1, 1, 1, QtCore.Qt.AlignHCenter)
        self.lineEdit_3 = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit_3.setFont(font)
        self.lineEdit_3.setObjectName("lineEdit_3")
        self.gridLayout.addWidget(self.lineEdit_3, 1, 2, 1, 1, QtCore.Qt.AlignHCenter)
        self.label_46 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_46.setFont(font)
        self.label_46.setObjectName("label_46")
        self.gridLayout.addWidget(self.label_46, 0, 3, 1, 1, QtCore.Qt.AlignHCenter)
        self.lineEdit = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit.setFont(font)
        self.lineEdit.setObjectName("lineEdit")
        self.gridLayout.addWidget(self.lineEdit, 0, 2, 1, 1, QtCore.Qt.AlignHCenter)
        self.label_3 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.gridLayout.addWidget(self.label_3, 2, 0, 1, 1)
        self.label_4 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")
        self.gridLayout.addWidget(self.label_4, 2, 1, 1, 1, QtCore.Qt.AlignHCenter)
        self.label_11 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_11.setFont(font)
        self.label_11.setObjectName("label_11")
        self.gridLayout.addWidget(self.label_11, 1, 0, 1, 1)
        self.label_5 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_5.setFont(font)
        self.label_5.setObjectName("label_5")
        self.gridLayout.addWidget(self.label_5, 0, 0, 1, 1)
        self.lineEdit_6 = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit_6.setFont(font)
        self.lineEdit_6.setObjectName("lineEdit_6")
        self.gridLayout.addWidget(self.lineEdit_6, 2, 2, 1, 1, QtCore.Qt.AlignHCenter)
        self.verticalLayout.addLayout(self.gridLayout)
        spacerItem = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.verticalLayout.addItem(spacerItem)
        self.pushButton = QtWidgets.QPushButton(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton.setFont(font)
        self.pushButton.setObjectName("pushButton")

        self.pushButton.clicked.connect(self.calculer_engrenage) # Liaison du clic du bouton à la fonction de calcul
        #################################################################################################

        self.verticalLayout.addWidget(self.pushButton, 0, QtCore.Qt.AlignHCenter)
        spacerItem1 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.verticalLayout.addItem(spacerItem1)
        self.pushButton_3 = QtWidgets.QPushButton(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_3.setFont(font)
        self.pushButton_3.setObjectName("pushButton_3")

        self.pushButton_3.clicked.connect(self.sauvegarder_dans_excel)  # Connecte le clic du bouton à la fonction de sauvegarde
        ########################################################################################
        
        self.verticalLayout.addWidget(self.pushButton_3, 0, QtCore.Qt.AlignHCenter)
        self.label_45 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_45.setFont(font)
        self.label_45.setObjectName("label_45")
        self.verticalLayout.addWidget(self.label_45, 0, QtCore.Qt.AlignHCenter)
        self.horizontalLayout_3 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_3.setObjectName("horizontalLayout_3")
        self.label_7 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_7.setFont(font)
        self.label_7.setObjectName("label_7")
        self.horizontalLayout_3.addWidget(self.label_7, 0, QtCore.Qt.AlignHCenter)
        self.label_8 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_8.setFont(font)
        self.label_8.setObjectName("label_8")
        self.horizontalLayout_3.addWidget(self.label_8, 0, QtCore.Qt.AlignHCenter)
        self.lineEdit_4 = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit_4.setFont(font)
        self.lineEdit_4.setObjectName("lineEdit_4")
        self.horizontalLayout_3.addWidget(self.lineEdit_4, 0, QtCore.Qt.AlignHCenter)
        self.label_9 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_9.setFont(font)
        self.label_9.setObjectName("label_9")
        self.horizontalLayout_3.addWidget(self.label_9, 0, QtCore.Qt.AlignHCenter)
        self.lineEdit_5 = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit_5.setFont(font)
        self.lineEdit_5.setObjectName("lineEdit_5")
        self.horizontalLayout_3.addWidget(self.lineEdit_5, 0, QtCore.Qt.AlignHCenter)
        self.verticalLayout.addLayout(self.horizontalLayout_3)
        spacerItem2 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.verticalLayout.addItem(spacerItem2)
        self.pushButton_2 = QtWidgets.QPushButton(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_2.setFont(font)
        self.pushButton_2.setObjectName("pushButton_2")
        self.verticalLayout.addWidget(self.pushButton_2, 0, QtCore.Qt.AlignHCenter)
        self.gridLayout_3.addLayout(self.verticalLayout, 0, 0, 1, 1)
        self.scrollArea.setWidget(self.scrollAreaWidgetContents)
        self.gridLayout_2.addWidget(self.scrollArea, 0, 0, 1, 1)

        self.retranslateUi(METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE)
        QtCore.QMetaObject.connectSlotsByName(METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE)

    def retranslateUi(self, METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE):
        _translate = QtCore.QCoreApplication.translate
        METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE.setWindowTitle(_translate("METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE", "METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE"))
        self.label.setText(_translate("METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE", "METHODE DE MAAG-TASCHENBUCH  POUR LA DENTURE HELICOIDALE"))
        self.label_2.setText(_translate("METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE", "Valeurs d\'entrée"))
        self.label_47.setText(_translate("METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE", "a\'"))
        self.label_31.setText(_translate("METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE", "Z1"))
        self.label_46.setText(_translate("METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE", "Z2"))
        self.label_3.setText(_translate("METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE", "Angle d\'hélice"))
        self.label_4.setText(_translate("METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE", "β"))
        self.label_11.setText(_translate("METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE", "Entraxe de fonctionnement"))
        self.label_5.setText(_translate("METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE", "Nombre de dent"))
        self.pushButton.setText(_translate("METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE", "CALCULEZ"))
        self.pushButton_3.setText(_translate("METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE", "SAUVEGARDER LES RESULTATS DU CALCUL"))
        self.label_45.setText(_translate("METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE", "Valeurs de sortie"))
        self.label_7.setText(_translate("METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE", "Les deports"))
        self.label_8.setText(_translate("METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE", "X1"))
        self.label_9.setText(_translate("METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE", "X2"))
        self.pushButton_2.setText(_translate("METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE", "RETOUR"))


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE = QtWidgets.QWidget()
    ui = Ui_METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE()
    ui.setupUi(METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE)
    METHODE_MAAG_TASCHENBUCH_D_HELICOIDALE.show()
    sys.exit(app.exec_())
