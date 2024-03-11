
import math
from PyQt5 import QtCore, QtGui, QtWidgets
import openpyxl
from PyQt5.QtWidgets import QMessageBox

class Ui_calcul_module(object):

    def calculer_et_afficher(self):
        try:
            T = float(self.lineEdit.text())
            k = float(self.lineEdit1.text())
            Rpe = float(self.lineEdit2.text())
            beta = float(self.lineEdit3.text())
        except ValueError:
            QMessageBox.warning(None, "Erreur de saisie", "Il est impossible d'effectuer le calcul. Vérifiez les valeurs d'entrée.")
            return

        class CalculM:
            def __init__(self, T, k, Rpe, beta):
                self.T = T
                self.k = k
                self.Rpe = Rpe
                self.beta = beta

            def calcul_m(self):
                return 2.34 * ((self.T/math.cos(math.radians(self.beta))) / (self.k * self.Rpe)) ** 0.5

        class ValeurProche:
            @staticmethod
            def trouver_valeur_proche(m, liste):
                valeur_proche = min(liste, key=lambda x: abs(x - m))
                return valeur_proche

        calcul = CalculM(T, k, Rpe, beta)
        m = calcul.calcul_m()

        liste = [0.06, 0.08, 0.10, 0.12, 0.15, 0.20, 0.25, 0.30, 0.40, 0.50, 0.75, 1, 1.25, 1.5, 2, 2.5, 3, 4, 5, 6, 8, 10, 12, 16, 20, 25, 32, 40, 50, 60, 0.07, 0.09, 0.11, 0.14, 0.18, 0.22, 0.28, 0.35, 0.45, 0.55, 0.7, 0.9, 1.125, 1.375, 1.75, 2.75, 3.5, 4.5, 5.5, 7, 9, 11, 14, 18, 22, 28, 36, 45, 55, 70]

        if m in liste:
            self.lineEdit4.setText(str(m))
        else:
            valeur_proche = ValeurProche.trouver_valeur_proche(m, liste)
            self.lineEdit4.setText(str(valeur_proche))  

    def sauvegarder_dans_excel(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        noms_labels = [
            self.label.text(),
            self.label1.text(),
            self.label2.text(),
            self.label3.text(),
            self.label4.text()

        ]
        symboles = ['T', 'K', 'Rpe', 'β', 'm']
        valeurs = [
            self.lineEdit.text(),
            self.lineEdit1.text(),
            self.lineEdit2.text(),
            self.lineEdit3.text(),
            self.lineEdit4.text()
        ]
        unites = ['N', '', 'N/mm', 'degré', 'mm']

        entete = ['Nom', 'Symbole', 'Valeur', 'Unité']
        ws.append(entete)

        for i in range(len(noms_labels)):
            ws.append([noms_labels[i], symboles[i], valeurs[i], unites[i]])

        for row in ws.iter_rows(min_row=1, max_row=6, min_col=1, max_col=4):
            for cell in row:
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(border_style='thin'),
                    right=openpyxl.styles.Side(border_style='thin'),
                    top=openpyxl.styles.Side(border_style='thin'),
                    bottom=openpyxl.styles.Side(border_style='thin')
                )

        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=4):
            for cell in row:
                cell.fill = openpyxl.styles.PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
                cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)


        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(None, "Enregistrer le fichier Excel", "", "Excel Files (*.xlsx)")
        if file_path:
            wb.save(file_path)
            QtWidgets.QMessageBox.information(None, "Enregistrement", "Fichier Excel enregistré avec succès !")

    def setupUi(self, calcul_module):
        calcul_module.setObjectName("calcul_module")
        calcul_module.resize(645, 480)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("../PFE/INTERFACE/biblio/ui/reducteur_gearexpert.gif"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        calcul_module.setWindowIcon(icon)
        self.gridLayout_2 = QtWidgets.QGridLayout(calcul_module)
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.scrollArea = QtWidgets.QScrollArea(calcul_module)
        self.scrollArea.setWidgetResizable(True)
        self.scrollArea.setObjectName("scrollArea")
        self.scrollAreaWidgetContents = QtWidgets.QWidget()
        self.scrollAreaWidgetContents.setGeometry(QtCore.QRect(0, 0, 631, 466))
        self.scrollAreaWidgetContents.setObjectName("scrollAreaWidgetContents")
        self.gridLayout_3 = QtWidgets.QGridLayout(self.scrollAreaWidgetContents)
        self.gridLayout_3.setObjectName("gridLayout_3")
        self.verticalLayout = QtWidgets.QVBoxLayout()
        self.verticalLayout.setObjectName("verticalLayout")
        self.label_2 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.verticalLayout.addWidget(self.label_2, 0, QtCore.Qt.AlignHCenter)
        self.gridLayout = QtWidgets.QGridLayout()
        self.gridLayout.setObjectName("gridLayout")
        self.label_48 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_48.setFont(font)
        self.label_48.setObjectName("label_48")
        self.gridLayout.addWidget(self.label_48, 2, 1, 1, 1, QtCore.Qt.AlignHCenter)
        self.label2 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label2.setFont(font)
        self.label2.setObjectName("label2")
        self.gridLayout.addWidget(self.label2, 2, 0, 1, 1)
        self.label_47 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_47.setFont(font)
        self.label_47.setObjectName("label_47")
        self.gridLayout.addWidget(self.label_47, 1, 1, 1, 1, QtCore.Qt.AlignHCenter)
        self.label1 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label1.setFont(font)
        self.label1.setObjectName("label1")
        self.gridLayout.addWidget(self.label1, 1, 0, 1, 1)
        self.label_31 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_31.setFont(font)
        self.label_31.setObjectName("label_31")
        self.gridLayout.addWidget(self.label_31, 0, 1, 1, 1, QtCore.Qt.AlignHCenter)
        self.lineEdit1 = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit1.setFont(font)
        self.lineEdit1.setObjectName("lineEdit1")
        self.gridLayout.addWidget(self.lineEdit1, 1, 2, 1, 1, QtCore.Qt.AlignHCenter)
        self.label = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.gridLayout.addWidget(self.label, 0, 0, 1, 1)
        self.lineEdit3 = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit3.setFont(font)
        self.lineEdit3.setObjectName("lineEdit3")
        self.gridLayout.addWidget(self.lineEdit3, 3, 2, 1, 1, QtCore.Qt.AlignHCenter)
        self.lineEdit = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit.setFont(font)
        self.lineEdit.setObjectName("lineEdit")
        self.gridLayout.addWidget(self.lineEdit, 0, 2, 1, 1, QtCore.Qt.AlignHCenter)
        self.label3 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label3.setFont(font)
        self.label3.setObjectName("label3")
        self.gridLayout.addWidget(self.label3, 3, 0, 1, 1)
        self.label_4 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")
        self.gridLayout.addWidget(self.label_4, 3, 1, 1, 1, QtCore.Qt.AlignHCenter)
        self.lineEdit2 = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit2.setFont(font)
        self.lineEdit2.setObjectName("lineEdit2")
        self.gridLayout.addWidget(self.lineEdit2, 2, 2, 1, 1, QtCore.Qt.AlignHCenter)
        self.verticalLayout.addLayout(self.gridLayout)
        self.pushButton_1 = QtWidgets.QPushButton(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_1.setFont(font)
        self.pushButton_1.setObjectName("pushButton_1")
        self.verticalLayout.addWidget(self.pushButton_1, 0, QtCore.Qt.AlignHCenter)
        spacerItem = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.verticalLayout.addItem(spacerItem)
        self.pushButton = QtWidgets.QPushButton(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton.setFont(font)
        self.pushButton.setObjectName("pushButton")

        self.pushButton.clicked.connect(self.calculer_et_afficher)  # Connecter le clic du bouton à la méthode de calcul

        self.verticalLayout.addWidget(self.pushButton, 0, QtCore.Qt.AlignHCenter)
        spacerItem1 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.verticalLayout.addItem(spacerItem1)
        self.pushButton_3 = QtWidgets.QPushButton(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_3.setFont(font)
        self.pushButton_3.setObjectName("pushButton_3")

        self.pushButton_3.clicked.connect(self.sauvegarder_dans_excel)  # Connecte le clic du bouton à la fonction de sauvegarde

        self.verticalLayout.addWidget(self.pushButton_3, 0, QtCore.Qt.AlignHCenter)
        self.label_45 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_45.setFont(font)
        self.label_45.setObjectName("label_45")
        self.verticalLayout.addWidget(self.label_45, 0, QtCore.Qt.AlignHCenter)
        self.horizontalLayout_3 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_3.setObjectName("horizontalLayout_3")
        self.label4 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label4.setFont(font)
        self.label4.setObjectName("label4")
        self.horizontalLayout_3.addWidget(self.label4, 0, QtCore.Qt.AlignLeft)
        self.label_8 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_8.setFont(font)
        self.label_8.setObjectName("label_8")
        self.horizontalLayout_3.addWidget(self.label_8, 0, QtCore.Qt.AlignHCenter)
        self.lineEdit4 = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit4.setFont(font)
        self.lineEdit4.setObjectName("lineEdit4")
        self.horizontalLayout_3.addWidget(self.lineEdit4, 0, QtCore.Qt.AlignHCenter)
        self.verticalLayout.addLayout(self.horizontalLayout_3)
        spacerItem2 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.verticalLayout.addItem(spacerItem2)
        self.bouton_retour_ch13 = QtWidgets.QPushButton(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.bouton_retour_ch13.setFont(font)
        self.bouton_retour_ch13.setObjectName("bouton_retour_ch13")
        self.verticalLayout.addWidget(self.bouton_retour_ch13, 0, QtCore.Qt.AlignHCenter)
        self.gridLayout_3.addLayout(self.verticalLayout, 0, 0, 1, 1)
        self.scrollArea.setWidget(self.scrollAreaWidgetContents)
        self.gridLayout_2.addWidget(self.scrollArea, 0, 0, 1, 1)

        self.label_I = QtWidgets.QLabel(calcul_module)
        self.label_I.setGeometry(QtCore.QRect(100, 200, 500, 200))
        self.label_I.setScaledContents(True)
        self.label_I.setObjectName("label_I")
        self.label_I.hide()  # Permet de cacher l'image au démarrage

        self.image_paths = {
            self.pushButton_1: "C:/Users/HD/Desktop/PFE/INTERFACE/biblio/module.png"
        }

        self.retranslateUi(calcul_module)
        QtCore.QMetaObject.connectSlotsByName(calcul_module)

    def retranslateUi(self, calcul_module):
        _translate = QtCore.QCoreApplication.translate
        calcul_module.setWindowTitle(_translate("calcul_module", "CALCUL_DU_MODULE"))
        self.label_2.setText(_translate("calcul_module", "Valeurs d\'entrée"))
        self.label_48.setText(_translate("calcul_module", "Rpe"))
        self.label2.setText(_translate("calcul_module", "Résistance pratique à l\'extension "))
        self.label_47.setText(_translate("calcul_module", "K"))
        self.label1.setText(_translate("calcul_module", "Coefficient de largeur de denture "))
        self.label_31.setText(_translate("calcul_module", "T"))
        self.label.setText(_translate("calcul_module", "Effort tangentiel sur la dent "))
        self.label3.setText(_translate("calcul_module", "Angle d\'hélice"))
        self.label_4.setText(_translate("calcul_module", "β"))
        self.pushButton_1.setText(_translate("calcul_module", "Aide"))
        self.pushButton.setText(_translate("calcul_module", "CALCULEZ"))
        self.pushButton_3.setText(_translate("calcul_module", "SAUVEGARDER LES RESULTATS DU CALCUL"))
        self.label_45.setText(_translate("calcul_module", "Valeurs de sortie"))
        self.label4.setText(_translate("calcul_module", "La valeur normalisée du module "))
        self.label_8.setText(_translate("calcul_module", "m"))
        self.bouton_retour_ch13.setText(_translate("calcul_module", "RETOUR"))

        self.pushButton_1.clicked.connect(lambda: self.toggleImage(self.pushButton_1))

    def toggleImage(self, button):
        if self.label_I.isHidden():
            image_path = self.image_paths[button]
            self.label_I.setPixmap(QtGui.QPixmap(image_path))
            self.label_I.show()
        else:
            self.label_I.hide()    

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    calcul_module = QtWidgets.QWidget()
    ui = Ui_calcul_module()
    ui.setupUi(calcul_module)
    calcul_module.show()
    sys.exit(app.exec_())

