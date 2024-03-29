
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMessageBox
import openpyxl
import math


class Ui_LES_EFFORTS_ROUE_ET_VIS(object):

    def perform_selected_calculation(self):

        R = ""
        if self.radioButton.isChecked():
            R = "vis"
        elif self.radioButton_2.isChecked():
            R = "roue"
        else:
            QMessageBox.warning(None, "Erreur de sélection", "Veuillez sélectionner entre la vis et la roue celle qui est menante.")
            return

        try:
            
            P = float(self.lineEdit1.text())
            N = float(self.lineEdit2.text())
            beta = float(self.lineEdit2_1.text())
            alpha = float(self.lineEdit2_2.text())
            D = float(self.lineEdit3.text())
            f = float(self.lineEdit4.text())  

            if R == "vis":

                C1 = (30*P)/(math.pi*N)
                FT1 = (C1/(D/2))*1000
                F1 = FT1/(math.cos(math.radians(alpha))*math.sin(math.radians(beta))+f*math.cos(math.radians(beta)))
                FA1 = F1*(math.cos(math.radians(alpha))*math.cos(math.radians(beta))-f*math.sin(math.radians(beta)))
                FR1 = F1*math.sin(math.radians(alpha))
    

                C2 = C1
                FT2 = FA1
                FA2 = FT1
                FR2 = FR1
                F2 = (FT2**2 + FA2**2 + FR2**2)**0.5
                η = ((math.cos(math.radians(alpha))-f*math.tan(math.radians(beta)))/(math.cos(math.radians(alpha))+f*(1/math.tan(math.radians(beta)))))*100
                            

            elif R == "roue":
                C2 = (30*P)/(math.pi*N)
                FT2 = (C2/(D/2))*1000
                F2 = FT2/(math.cos(math.radians(alpha))*math.sin(math.radians(beta))+f*math.cos(math.radians(beta)))
                FA2 = F2*(math.cos(math.radians(alpha))*math.cos(math.radians(beta))-f*math.sin(math.radians(beta)))
                FR2 = F2*math.sin(math.radians(alpha))
                
                C1 = C2
                FT1 = FA2
                FA1 = FT2
                FR1 = FR2
                F1 = (FT1**2 + FA1**2 + FR1**2)**0.5
                η = ((math.cos(math.radians(alpha))-f*math.tan(math.radians(beta)))/(math.cos(math.radians(alpha))+f*math.tan(math.radians(beta))))*100

                
        except ValueError:
            QMessageBox.warning(None, "Erreur de saisie", "Il est impossible d'effectuer le calcul. Vérifiez les valeurs d'entrée.")
            return
        except ZeroDivisionError:
            QMessageBox.warning(None, "Erreur de saisie", "Il est impossible d'effectuer le calcul. Vérifiez les valeurs d'entrée.")
            return 
        except UnboundLocalError:  
            QMessageBox.warning(None, "Erreur de sélection", "Il est impossible d'effectuer le calcul. sélectionez la roue menante.")
            return


        # Affichage des resultats dans les LineEdit
        self.lineEdit7.setText(str(C1))
        self.lineEdit8.setText(str(FT1))
        self.lineEdit8_1.setText(str(FA1))
        self.lineEdit9.setText(str(FR1))
        self.lineEdit10.setText(str(F1))
        self.lineEdit11.setText(str(C2))
        self.lineEdit12.setText(str(FT2))
        self.lineEdit12_1.setText(str(FA2))
        self.lineEdit13.setText(str(FR2))
        self.lineEdit14.setText(str(F2)) 
        self.lineEdit14_2.setText(str(η))   


    def sauvegarder_dans_excel(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        noms_labels = [
            self.label1.text(),
            self.label2.text(),
            self.label2_1.text(),
            self.label2_2.text(),
            self.label3.text(),
            self.label4.text(),
            self.label7.text(),
            self.label8.text(),
            self.label8_1.text(),
            self.label9.text(),
            self.label10.text(),
            self.label11.text(),
            self.label12.text(),
            self.label12_1.text(),
            self.label13.text(),
            self.label14.text(),
            self.label5_2.text(),
            
        ]

        symboles = [
            'P', 'N', 'βr', '∝', 'D', 'f', 'C1', 'FT1', 'FA1', 'FR1', 'F1', 'C2', 'FT2', 'FA2', 'FR2', 'F2', 'η'
        ]

        valeurs = [

            self.lineEdit1.text(),
            self.lineEdit2.text(),
            self.lineEdit2_1.text(),
            self.lineEdit2_2.text(),
            self.lineEdit3.text(),
            self.lineEdit4.text(),   
            self.lineEdit7.text(),
            self.lineEdit8.text(),
            self.lineEdit8_1.text(),
            self.lineEdit9.text(),
            self.lineEdit10.text(),
            self.lineEdit11.text(),
            self.lineEdit12.text(),
            self.lineEdit12_1.text(),
            self.lineEdit13.text(),
            self.lineEdit14.text(),
            self.lineEdit14_2.text(),
   
        ]

        unites = [

            'W', 'tr/mn', 'degré', 'degré',  'mm', '', 'N.m', 'N', 'N', 'N', 'N', 'N.m', 'N', 'N', 'N', 'N', '%'
        ]

        entete = ['Nom(s)', 'Symbole(s)', 'Valeur(s)', 'Unité(s)']
        ws.append(entete)

        for i in range(len(noms_labels)):
            ws.append([noms_labels[i], symboles[i], valeurs[i], unites[i]])

        for row in ws.iter_rows(min_row=1, max_row=18, min_col=1, max_col=4):
            for cell in row:
                cell.border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(border_style='thin'),
                    right=openpyxl.styles.Side(border_style='thin'),
                    top=openpyxl.styles.Side(border_style='thin'),
                    bottom=openpyxl.styles.Side(border_style='thin')
                )

        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=4):
            for cell in row:
                cell.fill = openpyxl.styles.PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
                cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)


        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(None, "Enregistrer le fichier Excel", "", "Excel Files (*.xlsx)")
        if file_path:
            wb.save(file_path)
            QtWidgets.QMessageBox.information(None, "Enregistrement", "Fichier Excel enregistré avec succès !")







    def setupUi(self, LES_EFFORTS_ROUE_ET_VIS):
        LES_EFFORTS_ROUE_ET_VIS.setObjectName("LES_EFFORTS_ROUE_ET_VIS")
        LES_EFFORTS_ROUE_ET_VIS.setWindowModality(QtCore.Qt.WindowModal)
        LES_EFFORTS_ROUE_ET_VIS.resize(470, 399)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("../../../../../.designer/backup/reducteur_gearexpert.gif"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        LES_EFFORTS_ROUE_ET_VIS.setWindowIcon(icon)
        self.gridLayout = QtWidgets.QGridLayout(LES_EFFORTS_ROUE_ET_VIS)
        self.gridLayout.setObjectName("gridLayout")
        self.scrollArea = QtWidgets.QScrollArea(LES_EFFORTS_ROUE_ET_VIS)
        self.scrollArea.setWidgetResizable(True)
        self.scrollArea.setObjectName("scrollArea")
        self.scrollAreaWidgetContents = QtWidgets.QWidget()
        self.scrollAreaWidgetContents.setGeometry(QtCore.QRect(0, -643, 443, 1028))
        self.scrollAreaWidgetContents.setObjectName("scrollAreaWidgetContents")
        self.gridLayout_6 = QtWidgets.QGridLayout(self.scrollAreaWidgetContents)
        self.gridLayout_6.setObjectName("gridLayout_6")
        self.verticalLayout = QtWidgets.QVBoxLayout()
        self.verticalLayout.setObjectName("verticalLayout")
        self.label_2 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.verticalLayout.addWidget(self.label_2, 0, QtCore.Qt.AlignHCenter)
        spacerItem = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.verticalLayout.addItem(spacerItem)
        self.horizontalLayout = QtWidgets.QHBoxLayout()
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.label_7 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_7.setFont(font)
        self.label_7.setObjectName("label_7")
        self.horizontalLayout.addWidget(self.label_7, 0, QtCore.Qt.AlignRight)
        self.verticalLayout_2 = QtWidgets.QVBoxLayout()
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        self.radioButton_2 = QtWidgets.QRadioButton(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.radioButton_2.setFont(font)
        self.radioButton_2.setObjectName("radioButton_2")
        self.verticalLayout_2.addWidget(self.radioButton_2)
        self.radioButton = QtWidgets.QRadioButton(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.radioButton.setFont(font)
        self.radioButton.setObjectName("radioButton")
        self.verticalLayout_2.addWidget(self.radioButton)
        self.horizontalLayout.addLayout(self.verticalLayout_2)
        self.verticalLayout.addLayout(self.horizontalLayout)
        spacerItem1 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.verticalLayout.addItem(spacerItem1)
        self.gridLayout_2 = QtWidgets.QGridLayout()
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.label_52 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_52.setFont(font)
        self.label_52.setObjectName("label_52")
        self.gridLayout_2.addWidget(self.label_52, 2, 1, 1, 1, QtCore.Qt.AlignHCenter)
        self.label_154 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_154.setFont(font)
        self.label_154.setObjectName("label_154")
        self.gridLayout_2.addWidget(self.label_154, 4, 1, 1, 1, QtCore.Qt.AlignHCenter)
        self.label1 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label1.setFont(font)
        self.label1.setObjectName("label1")
        self.gridLayout_2.addWidget(self.label1, 0, 0, 1, 1, QtCore.Qt.AlignLeft)
        self.label3 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label3.setFont(font)
        self.label3.setObjectName("label3")
        self.gridLayout_2.addWidget(self.label3, 4, 0, 1, 1, QtCore.Qt.AlignLeft)
        self.lineEdit1 = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit1.setFont(font)
        self.lineEdit1.setObjectName("lineEdit1")
        self.gridLayout_2.addWidget(self.lineEdit1, 0, 3, 1, 1, QtCore.Qt.AlignHCenter)
        self.label_27 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_27.setFont(font)
        self.label_27.setObjectName("label_27")
        self.gridLayout_2.addWidget(self.label_27, 1, 1, 1, 1, QtCore.Qt.AlignHCenter)
        self.label2 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label2.setFont(font)
        self.label2.setObjectName("label2")
        self.gridLayout_2.addWidget(self.label2, 1, 0, 1, 1, QtCore.Qt.AlignLeft)
        self.lineEdit4 = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit4.setFont(font)
        self.lineEdit4.setObjectName("lineEdit4")
        self.gridLayout_2.addWidget(self.lineEdit4, 5, 3, 1, 1, QtCore.Qt.AlignHCenter)
        self.lineEdit2 = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit2.setFont(font)
        self.lineEdit2.setObjectName("lineEdit2")
        self.gridLayout_2.addWidget(self.lineEdit2, 1, 3, 1, 1, QtCore.Qt.AlignHCenter)
        self.label_25 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_25.setFont(font)
        self.label_25.setObjectName("label_25")
        self.gridLayout_2.addWidget(self.label_25, 0, 1, 1, 1, QtCore.Qt.AlignHCenter)
        self.label4 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label4.setFont(font)
        self.label4.setObjectName("label4")
        self.gridLayout_2.addWidget(self.label4, 5, 0, 1, 1, QtCore.Qt.AlignLeft)
        self.label_170 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_170.setFont(font)
        self.label_170.setObjectName("label_170")
        self.gridLayout_2.addWidget(self.label_170, 5, 1, 1, 1, QtCore.Qt.AlignHCenter)
        self.lineEdit3 = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit3.setFont(font)
        self.lineEdit3.setObjectName("lineEdit3")
        self.gridLayout_2.addWidget(self.lineEdit3, 4, 3, 1, 1, QtCore.Qt.AlignHCenter)
        self.lineEdit2_2 = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit2_2.setFont(font)
        self.lineEdit2_2.setObjectName("lineEdit2_2")
        self.gridLayout_2.addWidget(self.lineEdit2_2, 3, 3, 1, 1, QtCore.Qt.AlignHCenter)
        self.label_54 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_54.setFont(font)
        self.label_54.setObjectName("label_54")
        self.gridLayout_2.addWidget(self.label_54, 3, 1, 1, 1, QtCore.Qt.AlignHCenter)
        self.label2_2 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label2_2.setFont(font)
        self.label2_2.setObjectName("label2_2")
        self.gridLayout_2.addWidget(self.label2_2, 3, 0, 1, 1)
        self.lineEdit2_1 = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit2_1.setFont(font)
        self.lineEdit2_1.setObjectName("lineEdit2_1")
        self.gridLayout_2.addWidget(self.lineEdit2_1, 2, 3, 1, 1, QtCore.Qt.AlignHCenter)
        self.label2_1 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label2_1.setFont(font)
        self.label2_1.setObjectName("label2_1")
        self.gridLayout_2.addWidget(self.label2_1, 2, 0, 1, 1, QtCore.Qt.AlignLeft)
        self.verticalLayout.addLayout(self.gridLayout_2)
        spacerItem2 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.verticalLayout.addItem(spacerItem2)
        self.pushButton = QtWidgets.QPushButton(self.scrollAreaWidgetContents)
        self.pushButton.setMaximumSize(QtCore.QSize(180, 180))
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton.setFont(font)
        self.pushButton.setObjectName("pushButton")

        self.pushButton.clicked.connect(self.perform_selected_calculation) # Liaison du clic du bouton à la fonction de calcul
        ##################################################################################################################

        self.verticalLayout.addWidget(self.pushButton, 0, QtCore.Qt.AlignHCenter)
        spacerItem3 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.verticalLayout.addItem(spacerItem3)
        self.pushButton_3 = QtWidgets.QPushButton(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_3.setFont(font)
        self.pushButton_3.setObjectName("pushButton_3")

        self.pushButton_3.clicked.connect(self.sauvegarder_dans_excel)  # Connecte le clic du bouton à la fonction de sauvegarde
        ###################################################################################################################

        self.verticalLayout.addWidget(self.pushButton_3, 0, QtCore.Qt.AlignHCenter)
        spacerItem4 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.verticalLayout.addItem(spacerItem4)
        self.label_45 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_45.setFont(font)
        self.label_45.setObjectName("label_45")
        self.verticalLayout.addWidget(self.label_45, 0, QtCore.Qt.AlignHCenter)
        spacerItem5 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.verticalLayout.addItem(spacerItem5)
        self.label_46 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_46.setFont(font)
        self.label_46.setObjectName("label_46")
        self.verticalLayout.addWidget(self.label_46, 0, QtCore.Qt.AlignHCenter)
        spacerItem6 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.verticalLayout.addItem(spacerItem6)
        self.gridLayout_3 = QtWidgets.QGridLayout()
        self.gridLayout_3.setObjectName("gridLayout_3")
        self.label_53 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_53.setFont(font)
        self.label_53.setObjectName("label_53")
        self.gridLayout_3.addWidget(self.label_53, 4, 1, 1, 1, QtCore.Qt.AlignHCenter)
        self.lineEdit9 = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit9.setFont(font)
        self.lineEdit9.setObjectName("lineEdit9")
        self.gridLayout_3.addWidget(self.lineEdit9, 3, 2, 1, 1, QtCore.Qt.AlignHCenter)
        self.lineEdit7 = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit7.setFont(font)
        self.lineEdit7.setObjectName("lineEdit7")
        self.gridLayout_3.addWidget(self.lineEdit7, 0, 2, 1, 1, QtCore.Qt.AlignHCenter)
        self.label9 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label9.setFont(font)
        self.label9.setObjectName("label9")
        self.gridLayout_3.addWidget(self.label9, 3, 0, 1, 1)
        self.label10 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label10.setFont(font)
        self.label10.setObjectName("label10")
        self.gridLayout_3.addWidget(self.label10, 4, 0, 1, 1)
        self.label_60 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_60.setFont(font)
        self.label_60.setObjectName("label_60")
        self.gridLayout_3.addWidget(self.label_60, 3, 1, 1, 1, QtCore.Qt.AlignHCenter)
        self.label_82 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_82.setFont(font)
        self.label_82.setObjectName("label_82")
        self.gridLayout_3.addWidget(self.label_82, 0, 1, 1, 1, QtCore.Qt.AlignHCenter)
        self.lineEdit10 = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit10.setFont(font)
        self.lineEdit10.setObjectName("lineEdit10")
        self.gridLayout_3.addWidget(self.lineEdit10, 4, 2, 1, 1, QtCore.Qt.AlignHCenter)
        self.label7 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label7.setFont(font)
        self.label7.setObjectName("label7")
        self.gridLayout_3.addWidget(self.label7, 0, 0, 1, 1)
        self.lineEdit8 = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit8.setFont(font)
        self.lineEdit8.setObjectName("lineEdit8")
        self.gridLayout_3.addWidget(self.lineEdit8, 1, 2, 1, 1, QtCore.Qt.AlignHCenter)
        self.label_83 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_83.setFont(font)
        self.label_83.setObjectName("label_83")
        self.gridLayout_3.addWidget(self.label_83, 1, 1, 1, 1, QtCore.Qt.AlignHCenter)
        self.label8 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label8.setFont(font)
        self.label8.setObjectName("label8")
        self.gridLayout_3.addWidget(self.label8, 1, 0, 1, 1)
        self.label8_1 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label8_1.setFont(font)
        self.label8_1.setObjectName("label8_1")
        self.gridLayout_3.addWidget(self.label8_1, 2, 0, 1, 1)
        self.label_88 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_88.setFont(font)
        self.label_88.setObjectName("label_88")
        self.gridLayout_3.addWidget(self.label_88, 2, 1, 1, 1, QtCore.Qt.AlignHCenter)
        self.lineEdit8_1 = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit8_1.setFont(font)
        self.lineEdit8_1.setObjectName("lineEdit8_1")
        self.gridLayout_3.addWidget(self.lineEdit8_1, 2, 2, 1, 1, QtCore.Qt.AlignHCenter)
        self.verticalLayout.addLayout(self.gridLayout_3)
        spacerItem7 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.verticalLayout.addItem(spacerItem7)
        self.label_50 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_50.setFont(font)
        self.label_50.setObjectName("label_50")
        self.verticalLayout.addWidget(self.label_50, 0, QtCore.Qt.AlignHCenter)
        self.gridLayout_5 = QtWidgets.QGridLayout()
        self.gridLayout_5.setObjectName("gridLayout_5")
        self.label12 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label12.setFont(font)
        self.label12.setObjectName("label12")
        self.gridLayout_5.addWidget(self.label12, 1, 0, 1, 1)
        self.label_87 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_87.setFont(font)
        self.label_87.setObjectName("label_87")
        self.gridLayout_5.addWidget(self.label_87, 1, 1, 1, 1, QtCore.Qt.AlignHCenter)
        self.lineEdit12 = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit12.setFont(font)
        self.lineEdit12.setObjectName("lineEdit12")
        self.gridLayout_5.addWidget(self.lineEdit12, 1, 2, 1, 1, QtCore.Qt.AlignHCenter)
        self.lineEdit14 = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit14.setFont(font)
        self.lineEdit14.setObjectName("lineEdit14")
        self.gridLayout_5.addWidget(self.lineEdit14, 4, 2, 1, 1, QtCore.Qt.AlignHCenter)
        self.label11 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label11.setFont(font)
        self.label11.setObjectName("label11")
        self.gridLayout_5.addWidget(self.label11, 0, 0, 1, 1)
        self.label_55 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_55.setFont(font)
        self.label_55.setObjectName("label_55")
        self.gridLayout_5.addWidget(self.label_55, 4, 1, 1, 1, QtCore.Qt.AlignHCenter)
        self.lineEdit13 = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit13.setFont(font)
        self.lineEdit13.setObjectName("lineEdit13")
        self.gridLayout_5.addWidget(self.lineEdit13, 3, 2, 1, 1, QtCore.Qt.AlignHCenter)
        self.lineEdit11 = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit11.setFont(font)
        self.lineEdit11.setObjectName("lineEdit11")
        self.gridLayout_5.addWidget(self.lineEdit11, 0, 2, 1, 1, QtCore.Qt.AlignHCenter)
        self.label14 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label14.setFont(font)
        self.label14.setObjectName("label14")
        self.gridLayout_5.addWidget(self.label14, 4, 0, 1, 1)
        self.label13 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label13.setFont(font)
        self.label13.setObjectName("label13")
        self.gridLayout_5.addWidget(self.label13, 3, 0, 1, 1)
        self.label_62 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_62.setFont(font)
        self.label_62.setObjectName("label_62")
        self.gridLayout_5.addWidget(self.label_62, 3, 1, 1, 1, QtCore.Qt.AlignHCenter)
        self.label_86 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_86.setFont(font)
        self.label_86.setObjectName("label_86")
        self.gridLayout_5.addWidget(self.label_86, 0, 1, 1, 1, QtCore.Qt.AlignHCenter)
        self.label12_1 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label12_1.setFont(font)
        self.label12_1.setObjectName("label12_1")
        self.gridLayout_5.addWidget(self.label12_1, 2, 0, 1, 1)
        self.label_89 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_89.setFont(font)
        self.label_89.setObjectName("label_89")
        self.gridLayout_5.addWidget(self.label_89, 2, 1, 1, 1, QtCore.Qt.AlignHCenter)
        self.lineEdit12_1 = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit12_1.setFont(font)
        self.lineEdit12_1.setObjectName("lineEdit12_1")
        self.gridLayout_5.addWidget(self.lineEdit12_1, 2, 2, 1, 1, QtCore.Qt.AlignHCenter)
        self.verticalLayout.addLayout(self.gridLayout_5)
        spacerItem8 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.verticalLayout.addItem(spacerItem8)
        spacerItem9 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.verticalLayout.addItem(spacerItem9)
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout()
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        self.label5_2 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label5_2.setFont(font)
        self.label5_2.setObjectName("label5_2")
        self.horizontalLayout_2.addWidget(self.label5_2, 0, QtCore.Qt.AlignLeft)
        self.label_33 = QtWidgets.QLabel(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_33.setFont(font)
        self.label_33.setObjectName("label_33")
        self.horizontalLayout_2.addWidget(self.label_33, 0, QtCore.Qt.AlignHCenter)
        self.lineEdit14_2 = QtWidgets.QLineEdit(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit14_2.setFont(font)
        self.lineEdit14_2.setObjectName("lineEdit14_2")
        self.horizontalLayout_2.addWidget(self.lineEdit14_2, 0, QtCore.Qt.AlignHCenter)
        self.verticalLayout.addLayout(self.horizontalLayout_2)
        spacerItem10 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.verticalLayout.addItem(spacerItem10)
        spacerItem11 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        self.verticalLayout.addItem(spacerItem11)
        self.pushButton_2 = QtWidgets.QPushButton(self.scrollAreaWidgetContents)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_2.setFont(font)
        self.pushButton_2.setObjectName("pushButton_2")
        self.verticalLayout.addWidget(self.pushButton_2, 0, QtCore.Qt.AlignHCenter)
        self.gridLayout_6.addLayout(self.verticalLayout, 0, 0, 1, 1)
        self.scrollArea.setWidget(self.scrollAreaWidgetContents)
        self.gridLayout.addWidget(self.scrollArea, 0, 0, 1, 1)

        self.retranslateUi(LES_EFFORTS_ROUE_ET_VIS)
        QtCore.QMetaObject.connectSlotsByName(LES_EFFORTS_ROUE_ET_VIS)

    def retranslateUi(self, LES_EFFORTS_ROUE_ET_VIS):
        _translate = QtCore.QCoreApplication.translate
        LES_EFFORTS_ROUE_ET_VIS.setWindowTitle(_translate("LES_EFFORTS_ROUE_ET_VIS", "LES_EFFORTS_ROUE_ET_VIS"))
        self.label_2.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "Valeurs d\'entrée"))
        self.label_7.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "Veuillez choisir celle qui est la menante"))
        self.radioButton_2.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "Roue"))
        self.radioButton.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "Vis"))
        self.label_52.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "βr"))
        self.label_154.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "D"))
        self.label1.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "Puissance Vis/Roue"))
        self.label3.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "Diamètre Vis/Roue"))
        self.label_27.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "N"))
        self.label2.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "Nombre de tr/min Vis/Roue"))
        self.label_25.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "P"))
        self.label4.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "Coefficient de frottement"))
        self.label_170.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "f"))
        self.label_54.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "∝"))
        self.label2_2.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "Angle de pression reél"))
        self.label2_1.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "Angle d\'helice de la Roue"))
        self.pushButton.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "CALCULEZ"))
        self.pushButton_3.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "SAUVEGARDEZ LES RESULTATS DU CALCUL"))
        self.label_45.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "Valeurs de sortie"))
        self.label_46.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "Les efforts pour la vis"))
        self.label_53.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "F1"))
        self.label9.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "Effort radial"))
        self.label10.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "Effort normal"))
        self.label_60.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "FR1"))
        self.label_82.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "C1"))
        self.label7.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "Couple"))
        self.label_83.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "FT1"))
        self.label8.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "Effort tangentiel"))
        self.label8_1.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "Effort axial"))
        self.label_88.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "FA1"))
        self.label_50.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "Les efforts pour le roue"))
        self.label12.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "Effort tangentiel"))
        self.label_87.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "FT2"))
        self.label11.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "Couple"))
        self.label_55.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "F2"))
        self.label14.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "Effort normal"))
        self.label13.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "Effort radial"))
        self.label_62.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "FR2"))
        self.label_86.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "C2"))
        self.label12_1.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "Effort axial"))
        self.label_89.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "FA2"))
        self.label5_2.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "Rendement"))
        self.label_33.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "η"))
        self.pushButton_2.setText(_translate("LES_EFFORTS_ROUE_ET_VIS", "RETOUR"))


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    LES_EFFORTS_ROUE_ET_VIS = QtWidgets.QWidget()
    ui = Ui_LES_EFFORTS_ROUE_ET_VIS()
    ui.setupUi(LES_EFFORTS_ROUE_ET_VIS)
    LES_EFFORTS_ROUE_ET_VIS.show()
    sys.exit(app.exec_())
